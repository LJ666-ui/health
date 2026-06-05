"""
Excel 数据导出器
支持健康数据明细、AI分析报告、统计汇总等多种Excel格式
"""

import os
from typing import Dict, List, Optional, Any, Union
from datetime import datetime, timedelta
from dataclasses import dataclass
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side,
    NamedStyle, Protection
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter
from loguru import logger


@dataclass
class ExportConfig:
    """导出配置"""
    include_header: bool = True           # 包含表头
    auto_width: bool = True               # 自动调整列宽
    freeze_panes: bool = True             # 冻结首行
    add_filters: bool = True              # 添加筛选功能
    highlight_abnormal: bool = True       # 高亮异常数据
    password_protect: Optional[str] = None  # 密码保护（可选）
    sheet_name: str = "HealthData"        # 工作表名称
    max_rows_per_sheet: int = 1000000     # 每个Sheet最大行数（Excel限制）


class ExcelExporter:
    """Excel文件导出引擎"""
    
    # 预定义样式
    STYLES = {
        'header': {
            'font': Font(bold=True, color='FFFFFF', size=11),
            'fill': PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        },
        'normal': {
            'alignment': Alignment(vertical='center'),
            'border': Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
        },
        'abnormal': {
            'font': Font(color='CC0000', bold=True),
            'fill': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        },
        'title': {
            'font': Font(bold=True, size=16, color='1F4E79'),
            'alignment': Alignment(horizontal='center')
        },
        'subtitle': {
            'font': Font(italic=True, size=10, color='666666'),
            'alignment': Alignment(horizontal='center')
        }
    }
    
    def __init__(self, config: Optional[ExportConfig] = None):
        self.config = config or ExportConfig()
    
    def export_health_data(self, 
                          df: pd.DataFrame,
                          output_path: Optional[str] = None) -> Union[str, BytesIO]:
        """
        导出健康数据明细表
        
        Args:
            df: 健康数据DataFrame
            output_path: 输出路径（如果为None则返回BytesIO）
            
        Returns:
            文件路径或BytesIO对象
        """
        wb = Workbook()
        ws = wb.active
        ws.title = self.config.sheet_name
        
        # 写入标题行
        if self.config.include_header and not df.empty:
            headers = list(df.columns)
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                self._apply_style(cell, self.STYLES['header'])
        
        # 写入数据行
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
            for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 应用基础样式
                self._apply_style(cell, self.STYLES['normal'])
                
                # 高亮异常值
                if (self.config.highlight_abnormal and 
                    col_name == 'is_abnormal' and 
                    value == 1):
                    for c in range(1, len(row_data) + 1):
                        abnormal_cell = ws.cell(row=row_idx, column=c)
                        self._apply_style(abnormal_cell, self.STYLES['abnormal'])
        
        # 自动调整列宽
        if self.config.auto_width:
            self._auto_adjust_column_width(ws, df)
        
        # 冻结首行
        if self.config.freeze_panes:
            ws.freeze_panes = 'A2'
        
        # 添加自动筛选
        if self.config.add_filters:
            ws.auto_filter.ref = ws.dimensions
        
        # 密码保护（可选）
        if self.config.password_protect:
            ws.protection.sheet(
                sheet=True, 
                password=self.config.password_protect
            )
        
        # 保存或返回内存对象
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            logger.info(f"Excel文件已保存: {output_path} ({len(df)} 行)")
            return output_path
        else:
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
    
    def export_ai_report(self,
                        analysis_results: Dict,
                        df: Optional[pd.DataFrame] = None,
                        output_path: Optional[str] = None) -> Union[str, BytesIO]:
        """
        导出AI分析报告（多Sheet格式）
        
        Sheet列表：
          1. Executive Summary - 执行摘要
          2. Risk Prediction - 风险预测详情
          3. Anomaly Detection - 异常检测结果
          4. User Segmentation - 用户分群
          5. Raw Data - 原始数据（可选）
        """
        wb = Workbook()
        
        # ===== Sheet 1: 执行摘要 =====
        summary_ws = wb.active
        summary_ws.title = "Executive Summary"
        
        self._write_report_header(summary_ws, "ZhiHealth AI 分析报告")
        self._write_summary_section(summary_ws, analysis_results)
        
        # ===== Sheet 2-4: 各模块详细结果 =====
        modules = analysis_results.get('modules', {})
        
        module_sheets = [
            ('risk_prediction', 'Risk Prediction', self._write_risk_details),
            ('anomaly_detection', 'Anomaly Detection', self._write_anomaly_details),
            ('user_segmentation', 'User Segmentation', self._write_segmentation_details),
        ]
        
        for module_key, sheet_title, write_func in module_sheets:
            if module_key in modules:
                module_ws = wb.create_sheet(title=sheet_title)
                write_func(module_ws, modules[module_key])
        
        # ===== Sheet 5: 原始数据 =====
        if df is not None and not df.empty:
            data_ws = wb.create_sheet(title="Raw Data")
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row, 1):
                    cell = data_ws.cell(row=r_idx+1, column=c_idx, value=value)
                    if r_idx == 0:
                        self._apply_style(cell, self.STYLES['header'])
        
        # 保存
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            logger.info(f"AI报告已导出: {output_path}")
            return output_path
        else:
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
    
    def export_statistics_dashboard(self,
                                   stats_dict: Dict,
                                   output_path: Optional[str] = None) -> Union[str, BytesIO]:
        """
        导出统计仪表板数据（带图表）
        
        包含：
          - 关键KPI指标卡片
          - 趋势折线图
          - 分类柱状图/饼图
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistics Dashboard"
        
        # 标题
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"ZhiHealth 统计报表 - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        self._apply_style(title_cell, self.STYLES['title'])
        
        current_row = 3
        
        # KPI指标区域
        kpis = stats_dict.get('kpis', [])
        if kpis:
            ws.cell(row=current_row, column=1, value="📊 关键绩效指标 (KPIs)")
            self._apply_style(ws.cell(row=current_row, column=1), {'font': Font(bold=True, size=14)})
            current_row += 2
            
            for i, kpi in enumerate(kpis[:8]):  # 最多显示8个KPI
                col = (i % 4) * 2 + 1
                row_offset = (i // 4) + current_row
                
                name_cell = ws.cell(row=row_offset, column=col, value=kpi.get('name', ''))
                value_cell = ws.cell(row=row_offset, column=col+1, value=kpi.get('value', ''))
                
                self._apply_style(name_cell, {'font': Font(bold=True, size=10)})
                self._apply_style(value_cell, {'font': Font(size=16, color='2F5496')})
            
            current_row += 4
        
        # 详细统计数据表格
        detail_data = stats_dict.get('details', [])
        if detail_data:
            detail_df = pd.DataFrame(detail_data)
            
            ws.cell(row=current_row, column=1, value="📋 详细统计数据")
            self._apply_style(ws.cell(row=current_row, column=1), {'font': Font(bold=True, size=14)})
            current_row += 2
            
            for r_idx, row in enumerate(dataframe_to_rows(detail_df, index=False, header=True)):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=current_row + r_idx, column=c_idx, value=value)
                    if r_idx == 0:
                        self._apply_style(cell, self.STYLES['header'])
        
        # 保存
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            return output_path
        else:
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer
    
    # ==================== 私有辅助方法 ====================
    
    def _apply_style(self, cell, style_dict: dict):
        """应用样式到单元格"""
        for attr, value in style_dict.items():
            setattr(cell, attr, value)
    
    def _auto_adjust_column_width(self, ws, df: pd.DataFrame):
        """根据内容自动调整列宽"""
        for col_idx, column in enumerate(df.columns, 1):
            max_length = len(str(column))
            
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = min(max_length + 2, 50)  # 最大50字符宽度
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    def _write_report_header(self, ws, title: str):
        """写入报告标题头"""
        ws.merge_cells('A1:F1')
        header_cell = ws['A1']
        header_cell.value = title
        self._apply_style(header_cell, self.STYLES['title'])
        
        ws.merge_cells('A2:F2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = (
            f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
            f"数据范围: 最近30天 | "
            f"版本: ZhiHealth v2.0"
        )
        self._apply_style(subtitle_cell, self.STYLES['subtitle'])
    
    def _write_summary_section(self, ws, results: Dict):
        """写入执行摘要部分"""
        data_summary = results.get('data_summary', {})
        modules = results.get('modules', {})
        
        start_row = 4
        
        # 基本统计
        ws.cell(row=start_row, column=1, value="📊 数据概览")
        self._apply_style(ws.cell(row=start_row, column=1), {'font': Font(bold=True, size=12)})
        
        summary_items = [
            ("总记录数", data_summary.get('total_records', 0)),
            ("用户数量", data_summary.get('users_count', 0)),
            ("时间范围", data_summary.get('date_range', 'N/A')),
            ("分析模块数", len(modules)),
        ]
        
        for idx, (label, value) in enumerate(summary_items, start=start_row+1):
            ws.cell(row=idx, column=1, value=label)
            ws.cell(row=idx, column=2, value=str(value))
        
        # 模块状态摘要
        module_start = start_row + len(summary_items) + 3
        ws.cell(row=module_start, column=1, value="🤖 AI模块执行状态")
        self._apply_style(ws.cell(row=module_start, column=1), {'font': Font(bold=True, size=12)})
        
        headers = ['模块名称', '状态', '关键指标']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=module_start+1, column=col, value=h)
            self._apply_style(cell, self.STYLES['header'])
        
        for mod_idx, (mod_name, mod_result) in enumerate(modules.items()):
            row = module_start + 2 + mod_idx
            status = "✅ 完成" if mod_result else "⚠️ 未完成"
            
            ws.cell(row=row, column=1, value=mod_name.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=status)
            
            # 提取关键指标
            key_metric = self._extract_key_metric(mod_result)
            ws.cell(row=row, column=3, value=key_metric)
    
    def _extract_key_metric(self, mod_result: Dict) -> str:
        """从模块结果中提取关键指标文本"""
        metrics = mod_result.get('metrics', {})
        
        if 'accuracy' in metrics:
            return f"准确率: {metrics['accuracy']:.1%}"
        elif 'anomaly_rate' in metrics:
            return f"异常率: {metrics['anomaly_rate']:.2%}"
        elif 'silhouette_score' in metrics:
            return f"轮廓系数: {metrics['silhouette_score']:.3f}"
        elif 'clusters' in metrics:
            return f"分群数: {metrics['clusters']}"
        else:
            return str(list(metrics.keys())[:3])
    
    def _write_risk_details(self, ws, risk_results: Dict):
        """写入风险预测详情"""
        predictions = risk_results.get('predictions', [])
        metrics = risk_results.get('metrics', {})
        
        ws['A1'] = "🔴 健康风险预测详细结果"
        self._apply_style(ws['A1'], {'font': Font(bold=True, size=14, color='CC0000')})
        
        if predictions:
            pred_df = pd.DataFrame(predictions)
            for r_idx, row in enumerate(dataframe_to_rows(pred_df, index=False, header=True)):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx+3, column=c_idx, value=value)
                    if r_idx == 0:
                        self._apply_style(cell, self.STYLES['header'])
                    
                    # 根据风险等级着色
                    if r_idx > 0 and c_idx == 3:  # risk_level列
                        if value == 'critical':
                            cell.fill = PatternFill(start_color='FF0000', fill_type='solid')
                            cell.font = Font(color='FFFFFF', bold=True)
                        elif value == 'high':
                            cell.fill = PatternFill(start_color='FF6600', fill_type='solid')
                        elif value == 'medium':
                            cell.fill = PatternFill(start_color='FFCC00', fill_type='solid')
    
    def _write_anomaly_details(self, ws, anomaly_results: Dict):
        """写入异常检测详情"""
        anomalies = anomaly_results.get('anomalies', [])
        stats = anomaly_results.get('statistics', {})
        
        ws['A1'] = "⚠️ 异常检测结果"
        self._apply_style(ws['A1'], {'font': Font(bold=True, size=14, color='990000')})
        
        if anomalies:
            anom_df = pd.DataFrame(anomalies)
            for r_idx, row in enumerate(dataframe_to_rows(anom_df, index=False, header=True)):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx+3, column=c_idx, value=value)
                    if r_idx == 0:
                        self._apply_style(cell, self.STYLES['header'])
    
    def _write_segmentation_details(self, ws, seg_results: Dict):
        """写入用户分群详情"""
        profiles = seg_results.get('cluster_profiles', {})
        
        ws['A1'] = "👥 用户群体画像"
        self._apply_style(ws['A1'], {'font': Font(bold=True, size=14, color='0066CC')})
        
        row = 3
        for cluster_id, profile in profiles.items():
            ws.cell(row=row, column=1, value=f"群体 {cluster_id}: {profile.get('archetype', 'Unknown')}")
            self._apply_style(ws.cell(row=row, column=1), {'font': Font(bold=True, size=12)})
            row += 1
            
            details = [
                ("用户数", profile.get('user_count', 0)),
                ("占比", f"{profile.get('percentage', 0):.1f}%"),
                ("平均年龄", profile.get('avg_age', 'N/A')),
                ("主要特征", ", ".join(profile.get('key_characteristics', []))),
            ]
            
            for label, value in details:
                ws.cell(row=row, column=1, value=f"  • {label}")
                ws.cell(row=row, column=2, value=str(value))
                row += 1
            
            row += 1  # 空行分隔